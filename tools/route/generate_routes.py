#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""巡回員別 効率化ルート生成 (1日7件)
入力: 自社巡回業務一覧表.xlsx の「一覧」シート
出力: 巡回員ごとに近いエリアでまとめた日別ルート表 (xlsx)
"""
import re, math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict, Counter

SRC = 'data.xlsx'
OUT = '巡回ルート_1日7件_巡回員別.xlsx'
PER_DAY = 7

# --- 区市 おおよその中心座標 (lat, lng) : 近接クラスタリング用 ---
WARD_XY = {
    '新宿区':(35.6938,139.7036),'港区':(35.6581,139.7516),'豊島区':(35.7289,139.7100),
    '世田谷区':(35.6464,139.6532),'品川区':(35.6092,139.7302),'渋谷区':(35.6640,139.6982),
    '大田区':(35.5614,139.7161),'文京区':(35.7080,139.7522),'中央区':(35.6706,139.7720),
    '台東区':(35.7126,139.7800),'目黒区':(35.6415,139.6982),'江東区':(35.6729,139.8172),
    '墨田区':(35.7107,139.8015),'北区':(35.7528,139.7336),'千代田区':(35.6940,139.7536),
    '板橋区':(35.7512,139.7093),'中野区':(35.7074,139.6638),'杉並区':(35.6995,139.6365),
    '足立区':(35.7750,139.8046),'荒川区':(35.7361,139.7833),'練馬区':(35.7357,139.6517),
    '葛飾区':(35.7434,139.8472),'江戸川区':(35.7067,139.8683),
    '横浜市':(35.4437,139.6380),'川崎市':(35.5308,139.7029),'さいたま市':(35.8617,139.6455),
    '川口市':(35.8078,139.7241),'府中市':(35.6717,139.4778),'武蔵野市':(35.7178,139.5661),
    '柏市':(35.8676,139.9758),'八王子市':(35.6664,139.3159),'船橋市':(35.6947,139.9825),
    '習志野市':(35.6809,140.0267),'三郷市':(35.8307,139.8721),'所沢市':(35.7998,139.4687),
    '小金井市':(35.6994,139.5031),'国立市':(35.6836,139.4413),'町田市':(35.5461,139.4386),
    '戸田市':(35.8172,139.6779),'市川市':(35.7219,139.9310),'三鷹市':(35.6835,139.5595),
    '新座市':(35.7935,139.5654),'調布市':(35.6506,139.5410),'千葉市':(35.6074,140.1065),
}
DEFAULT_XY = (35.69,139.70)

def parse_ward(addr):
    m = re.match(r'(?:東京都|神奈川県|埼玉県|千葉県)?(.+?[区市])', addr)
    return m.group(1) if m else '?'

def parse_town(addr):
    # 区市より後ろ (町名+丁目+番地)
    m = re.match(r'(?:東京都|神奈川県|埼玉県|千葉県)?.+?[区市郡](.*)', addr)
    rest = m.group(1) if m else addr
    # 町名 (数字の前まで)
    tm = re.match(r'([^0-9０-９]+)', rest)
    town = tm.group(1) if tm else rest
    # 丁目・番地の最初の数字
    nums = re.findall(r'\d+', rest.translate(str.maketrans('０１２３４５６７８９','0123456789')))
    chome = int(nums[0]) if nums else 0
    banchi = int(nums[1]) if len(nums) > 1 else 0
    return town, chome, banchi

def haversine(a, b):
    R=6371.0
    la1,lo1=math.radians(a[0]),math.radians(a[1])
    la2,lo2=math.radians(b[0]),math.radians(b[1])
    dla,dlo=la2-la1,lo2-lo1
    h=math.sin(dla/2)**2+math.cos(la1)*math.cos(la2)*math.sin(dlo/2)**2
    return 2*R*math.asin(math.sqrt(h))

def order_wards(wards):
    """区の集合を最近傍法で地理順に並べる。開始は最も北西の区。"""
    pts={w:WARD_XY.get(w,DEFAULT_XY) for w in wards}
    remaining=set(wards)
    # 開始: 緯度が最大(北)の区
    cur=max(remaining,key=lambda w:pts[w][0])
    seq=[cur]; remaining.discard(cur)
    while remaining:
        nxt=min(remaining,key=lambda w:haversine(pts[cur],pts[w]))
        seq.append(nxt); remaining.discard(nxt); cur=nxt
    return seq

# --- 読み込み ---
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['一覧']
# ヘッダ: 3行目
props=[]
for r in range(4, ws.max_row+1):
    num=ws.cell(r,2).value
    name=ws.cell(r,3).value
    addr=ws.cell(r,4).value
    junkai=ws.cell(r,12).value
    shubetsu=ws.cell(r,13).value
    biko=ws.cell(r,14).value
    auto=ws.cell(r,7).value
    key=ws.cell(r,8).value
    if not name and not addr and not num:
        continue
    j=(str(junkai).strip() if junkai else 'なし')
    props.append(dict(num=num,name=name,addr=str(addr or ''),j=j,
                      shubetsu=shubetsu,biko=biko,auto=auto,key=key))

# 巡回員ごと
by_j=defaultdict(list)
for p in props:
    by_j[p['j']].append(p)

# 出力順: 件数多い順、末尾に 管理員/CAN/なし
special=['管理員','CAN','なし']
main=[k for k in by_j if k not in special]
main.sort(key=lambda k:-len(by_j[k]))
order=main+[k for k in special if k in by_j]

# --- ルート最適化 ---
def build_route(items):
    for p in items:
        p['ward']=parse_ward(p['addr'])
        p['town'],p['chome'],p['banchi']=parse_town(p['addr'])
    wards=list({p['ward'] for p in items})
    wseq=order_wards(wards)
    wrank={w:i for i,w in enumerate(wseq)}
    items.sort(key=lambda p:(wrank[p['ward']],p['town'],p['chome'],p['banchi']))
    return items

# --- Excel 出力 ---
out = openpyxl.Workbook()
out.remove(out.active)

thin=Side(style='thin',color='D9D9D9')
border=Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_fill=PatternFill('solid',fgColor='4472C4')
hdr_font=Font(bold=True,color='FFFFFF',size=10)
day_fill=PatternFill('solid',fgColor='FCE4D6')
day_font=Font(bold=True,size=11,color='843C0C')
center=Alignment(horizontal='center',vertical='center')
left=Alignment(horizontal='left',vertical='center',wrap_text=True)

COLS=[('日',6),('順',5),('物件番号',10),('物件名',26),('住所',34),
      ('エリア',9),('オートロック',16),('緊急キー',18),('備考',18)]

# サマリーシート
sm=out.create_sheet('サマリー')
sm['A1']='巡回員別 効率化ルート (1日7件)';sm['A1'].font=Font(bold=True,size=14)
sm['A2']='※担当割り当ては元データの「巡回員」列を踏襲。近いエリアごとにまとめ7件/日で日別に区切り。'
sm['A2'].font=Font(size=9,color='808080')
srow=4
for c,(h,_) in enumerate([('巡回員',14),('担当件数',10),('必要日数',10),('主なエリア',40)],1):
    cell=sm.cell(srow,c,h);cell.fill=hdr_fill;cell.font=hdr_font;cell.border=border;cell.alignment=center
srow+=1

for j in order:
    items=build_route(list(by_j[j]))
    ndays=math.ceil(len(items)/PER_DAY)
    ward_ct=Counter(p['ward'] for p in items)
    main_areas='、'.join(f'{w}({n})' for w,n in ward_ct.most_common(5))
    # サマリー行
    for c,val in enumerate([j,len(items),ndays,main_areas],1):
        cell=sm.cell(srow,c,val);cell.border=border
        cell.alignment=left if c==4 else center
    srow+=1

    # 各巡回員シート
    title=re.sub(r'[\\/*?:\[\]]','',str(j))[:28] or '空欄'
    sh=out.create_sheet(title)
    sh['A1']=f'{j} 巡回ルート（1日{PER_DAY}件・全{len(items)}件・{ndays}日）'
    sh['A1'].font=Font(bold=True,size=13);sh.merge_cells('A1:I1')
    hr=3
    for c,(h,w) in enumerate(COLS,1):
        cell=sh.cell(hr,c,h);cell.fill=hdr_fill;cell.font=hdr_font;cell.border=border;cell.alignment=center
        sh.column_dimensions[openpyxl.utils.get_column_letter(c)].width=w
    sh.freeze_panes='A4'
    rr=hr+1
    for i,p in enumerate(items):
        day=i//PER_DAY+1
        seq=i%PER_DAY+1
        vals=[day if seq==1 else '',seq,p['num'],p['name'],p['addr'],p['ward'],
              p['auto'],p['key'],p['biko']]
        for c,val in enumerate(vals,1):
            cell=sh.cell(rr,c,val);cell.border=border
            cell.alignment=center if c in(1,2,3,6) else left
            if seq==1 and c==1:
                cell.fill=day_fill;cell.font=day_font
        # 日の区切りに薄い上罫線
        if seq==1 and i>0:
            for c in range(1,len(COLS)+1):
                cur=sh.cell(rr,c).border
                sh.cell(rr,c).border=Border(left=cur.left,right=cur.right,
                    top=Side(style='medium',color='F4B183'),bottom=cur.bottom)
        rr+=1

for c,w in enumerate([14,10,10,60],1):
    sm.column_dimensions[openpyxl.utils.get_column_letter(c)].width=w

out.save(OUT)
print('saved',OUT)
print('巡回員数(出力シート):',len(order))
for j in order:
    n=len(by_j[j]);print(f'  {j}: {n}件 -> {math.ceil(n/PER_DAY)}日')
